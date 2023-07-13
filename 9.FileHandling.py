"""
import os
print (os.getcwd()) #현재 디렉토리 확인

#현재 작업 디렉토리 변경
#os.chdir('/Users/munseokpark/Documents/python/source/basic')
"""


"""
#파일에 기록하기
try:
    file = open('./data/test.txt', 'w')
    # 파일 객체 정보 출력
    print(file)
    # 데이터를 한번에 기록
    file.write('Hello Python')
    file.write('\n\n')
    # 데이터를 줄 단위로 기록
    lines = ['안녕하세요', '반갑습니다.', '파이썬입니다.']
    file.write('\n'.join(lines))

    '''
    \n이 포함된 경우는
    file.write(''.join(lines))
    file.writelines(lines)
    가능
    '''

    print('기록 성공')
except Exception as e:
    print("파일 처리 중 예외 발생:", e)
finally:
    file.close()
"""


"""
#파일의 내용 읽어오기
try:
    file = open('./data/test.txt', 'r')
    #한꺼번에 전부 읽기
    content = file.read()
    print(content)
except Exception as e:
    print("파일 처리 중 예외 발생:", e)
finally:
    file.close()
print("==============")
#줄단위로 읽기
try:
    file = open('./data/test.txt', 'r')
    for line in file:
        print(line)
except Exception as e:
    print("파일 처리 중 예외 발생:", e)
finally:
    file.close()
print("==============")

try:
    file = open('./data/test.txt', 'r')
    lines = file.readlines()
    print(lines)
except Exception as e:
    print("파일 처리 중 예외 발생:", e)
finally:
    file.close()
"""


"""
#test.csv 읽기
with open('./data/test.csv', 'r') as f:
    #한꺼번에 전부 읽기
    content = f.read()
    print("==============")
    ar = content.split(',')
    print(type(ar))
    for imsi in ar:
        print(imsi)
"""

"""
#singer.csv 읽기
import csv
with open('./data/singer.csv', 'r', encoding='cp949') as f:
    rdr = csv.reader(f)
    for line in rdr:
        print(line)
"""

"""
#singer.csv 쓰기

import csv
with open('./data/singer.csv', 'a', encoding='cp949', newline='') as f:
    wr = csv.writer(f)
    wr.writerow([4, "최유정", "1999-11-12", "위키미키"])
    wr.writerow([5, "아이유", "1993-05-16", "솔로"])
"""

"""
with open('data/test.bin', 'wb') as f:
    f.write("안녕하세요".encode())

with open('data/test.bin', 'rb') as f:
    byteAr = f.read()
    # print(byteAr)
    print(byteAr.decode())
"""


"""
class Dto:
    def setNum(self, num):
        self.num = num
    def setName(self, name):
        self.name = name
    def getNum(self):
        return self.num
    def getName(self):
        return self.name
    def toString(self):
        return "{번호:" + str(self.num) + ",이름:" + self.name + "}"
data1 = Dto()
data1.setNum(1)
data1.setName("park")
data2 = Dto()
data2.setNum(2)
data2.setName("kim")
li = [data1, data2]


import pickle
try:
    with open('./data/test.txt', 'wb') as f:
        pickle.dump(li, f)
        f.close()
    with open('./data/test.txt', 'rb') as f:
        result = pickle.load(f)
        for temp in result:
            print(temp.toString())
        f.close()
except Exception as e:
    print("예외 발생:", e)
finally:
    f.close()
"""

"""
import zipfile
filelist = ["data/test.txt"]
with zipfile.ZipFile('data/test.zip', 'w', compression=zipfile.ZIP_BZIP2) as myzip:
    for temp in filelist:
        myzip.write(temp)
zipfile.ZipFile('data/test.zip').extractall()
"""

"""
import tarfile
filelist = ["./data/test.txt"]
with tarfile.open('data/test.tar.gz', 'w:gz') as mytar:
    for temp in filelist:
        mytar.add(temp)
tarfile.open('data/test.tar.gz').extractall()
"""


"""
pageviews = 0
f = None
with open('data/log.txt', 'r') as f:
    logs = f.readlines()
    for log in logs:
        log = log.split()
        status = log[8]
        if status == '200':
            pageviews += 1
    print('총 페이지뷰: [%d]' %pageviews)
"""


"""
visit_ip = []
f = None
with open('data/log.txt', 'r') as f:
    logs = f.readlines()
    for log in logs:
        log = log.split()
        ip = log[0]
        if ip not in visit_ip:
            visit_ip.append(ip)
    print('방문자수: [%d]' %len(visit_ip))
"""


"""
total_service = 0
with open('data/log.txt', 'r') as f:
    logs = f.readlines()
    for log in logs:
        log = log.split()
        servicebyte = log[9]
        if servicebyte.isdigit():
            total_service += int(servicebyte)
    total_service /= 1024
    print('총 서비스 용량: %dKB' %total_service)
"""

"""
services = {}
f = None
with open('data/log.txt', 'r') as f:
    logs = f.readlines()
    for log in logs:
        log = log.split()
        ip = log[0]
        servicebyte = log[9]
        if servicebyte.isdigit():
            servicebyte = int(servicebyte)
        else:
            servicebyte = 0
        if ip not in services:
            services[ip] = servicebyte
        else:
            services[ip] += servicebyte
    ret = sorted(services.items(), key=lambda x: x[1], reverse=True)
    print('사용자IP – 서비스용량')
    for ip, b in ret:
        print('[%s] – [%d]' %(ip, b))
"""


"""
#엑셀 파일 읽기 와 쓰기
import openpyxl

wb = None
try:
    # 엑셀파일 열기
    wb = openpyxl.load_workbook('data/score.xlsx')

    # 현재 Active Sheet 얻기
    ws = wb.active
    # ws = wb.get_sheet_by_name("Sheet1“)

    for r in ws.rows:
        row_index = r[0].row  # 행 인덱스
        if row_index == 1:
            ws.cell(row=row_index, column=5).value = '합계'
            continue
        name = r[0].value
        kor = r[1].value
        eng = r[2].value
        math = r[3].value
        sum = kor + eng + math
        # 합계 쓰기
        ws.cell(row=row_index, column=5).value = sum
        print(name, ':' , kor, eng, math, sum)
    # 엑셀 파일 저장
    wb.save("data/score2.xlsx")
except Exception as e:
    print("예외 발생:", e)
finally:
    wb.close()
"""


